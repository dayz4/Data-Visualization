import openpyxl
import json
import csv

with open('s1.json') as s1json, open('t1.json') as t1json, open('t2.json') as t2json, open('graph_mindsumo_json.js', 'w') as output1:
	s1 = json.load(s1json)
	t1 = json.load(t1json)
	t2 = json.load(t2json)

	sheet1 = openpyxl.load_workbook('t1-s1.xlsx')
	sheet1 = sheet1.get_sheet_by_name('t1-s1')

	sheet2 = openpyxl.load_workbook('t2-s1.xlsx')
	sheet2 = sheet2.get_sheet_by_name('t2-s1')

	t1s1 = []
	tcrosswalks = []
	scrosswalks = []
	for row in range (2, len(sheet1.rows)+1):
		#Find cell names
		su = 'A' + str(row)
		sl = 'B' + str(row)
		sn = 'C' + str(row)
		sg = 'D' + str(row)
		pc = 'E' + str(row)
		pl = 'F' + str(row)
		ou = 'G' + str(row)
		ol = 'H' + str(row)
		on = 'I' + str(row)
		og = 'J' + str(row)
		ri = 'K' + str(row)

		su = sheet1[su].value
		sl = sheet1[sl].value
		sn = sheet1[sn].value
		sg = sheet1[sg].value
		pc = sheet1[pc].value
		pl = sheet1[pl].value
		ou = sheet1[ou].value
		ol = sheet1[ol].value
		on = sheet1[on].value
		og = sheet1[og].value
		ri = sheet1[ri].value

		tcrosswalks.append(su)
		t1s1.append(su)
		t1s1.append(sl)
		t1s1.append(sn)
		t1s1.append(sg)
		t1s1.append(pc)
		t1s1.append(pl)
		scrosswalks.append(ou)
		t1s1.append(ou)
		t1s1.append(ol)
		t1s1.append(on)
		t1s1.append(og)
		t1s1.append(ri)

	i = 0
	for tcrosswalk in tcrosswalks:
		if 's1URI' in t1[tcrosswalk].keys():
			t1[tcrosswalk]['s1URI'].append(t1s1[i+6])
			t1[tcrosswalk]['s1Label'].append(t1s1[i+7])
			t1[tcrosswalk]['s1Notation'].append(t1s1[i+8])
			t1[tcrosswalk]['s1Grade'].append(t1s1[i+9])
		else:
			t1[tcrosswalk]['Label'] = t1s1[i+1]
			t1[tcrosswalk]['Notation'] = t1s1[i+2]
			t1[tcrosswalk]['Grade'] = t1s1[i+3]
			t1[tcrosswalk]['predicateCURIE'] = t1s1[i+4]
			t1[tcrosswalk]['predicateLabel'] = t1s1[i+5]
			t1[tcrosswalk]['s1URI'] = [t1s1[i+6]]
			t1[tcrosswalk]['s1Label'] = [t1s1[i+7]]
			t1[tcrosswalk]['s1Notation'] = [t1s1[i+8]]
			t1[tcrosswalk]['s1Grade'] = [t1s1[i+9]]
			t1[tcrosswalk]['relationshipID'] = t1s1[i+10]
			t1[tcrosswalk]['s1crosswalked'] = 'true'
		i = i+11

	i = 0;
	for scrosswalk in scrosswalks:
		if 't1URI' in s1[scrosswalk].keys():
			s1[scrosswalk]['t1URI'].append(t1s1[i])
			s1[scrosswalk]['t1Label'].append(t1s1[i+1])
			s1[scrosswalk]['t1Notation'].append(t1s1[i+2])
			s1[scrosswalk]['t1Grade'].append(t1s1[i+3])
		else:
			s1[scrosswalk]['t1URI'] = [t1s1[i]]
			s1[scrosswalk]['t1Label'] = [t1s1[i+1]]
			s1[scrosswalk]['t1Notation'] = [t1s1[i+2]]
			s1[scrosswalk]['t1Grade'] = [t1s1[i+3]]
			s1[scrosswalk]['predicateCURIE'] = t1s1[i+4]
			s1[scrosswalk]['predicateLabel'] = t1s1[i+5]
			s1[scrosswalk]['Label'] = t1s1[i+7]
			s1[scrosswalk]['Notation'] = t1s1[i+8]
			s1[scrosswalk]['Grade'] = t1s1[i+9]
			s1[scrosswalk]['relationshipID'] = t1s1[i+10]
			s1[scrosswalk]['t1crosswalked'] = 'true'
		i = i+11

	t2s1 = []
	tcrosswalks = []
	scrosswalks = []
	for row in range (2, len(sheet2.rows)+1):
		#Find cell names
		su = 'A' + str(row)
		sl = 'B' + str(row)
		sn = 'C' + str(row)
		sg = 'D' + str(row)
		pc = 'E' + str(row)
		pl = 'F' + str(row)
		ou = 'G' + str(row)
		ol = 'H' + str(row)
		on = 'I' + str(row)
		og = 'J' + str(row)
		ri = 'K' + str(row)

		su = sheet2[su].value
		sl = sheet2[sl].value
		sn = sheet2[sn].value
		sg = sheet2[sg].value
		pc = sheet2[pc].value
		pl = sheet2[pl].value
		ou = sheet2[ou].value
		ol = sheet2[ol].value
		on = sheet2[on].value
		og = sheet2[og].value
		ri = sheet2[ri].value

		tcrosswalks.append(su)
		t2s1.append(su)
		t2s1.append(sl)
		t2s1.append(sn)
		t2s1.append(sg)
		t2s1.append(pc)
		t2s1.append(pl)
		scrosswalks.append(ou)
		t2s1.append(ou)
		t2s1.append(ol)
		t2s1.append(on)
		t2s1.append(og)
		t2s1.append(ri)

	i = 0
	for tcrosswalk in tcrosswalks:
		if 's1URI' in t2[tcrosswalk].keys():
			t2[tcrosswalk]['s1URI'].append(t2s1[i+6])
			t2[tcrosswalk]['s1Label'].append(t2s1[i+7])
			t2[tcrosswalk]['s1Notation'].append(t2s1[i+8])
			t2[tcrosswalk]['s1Grade'].append(t2s1[i+9])
		else:
			t2[tcrosswalk]['Label'] = t2s1[i+1]
			t2[tcrosswalk]['Notation'] = t2s1[i+2]
			t2[tcrosswalk]['Grade'] = t2s1[i+3]
			t2[tcrosswalk]['predicateCURIE'] = t2s1[i+4]
			t2[tcrosswalk]['predicateLabel'] = t2s1[i+5]
			t2[tcrosswalk]['s1URI'] = [t2s1[i+6]]
			t2[tcrosswalk]['s1Label'] = [t2s1[i+7]]
			t2[tcrosswalk]['s1Notation'] = [t2s1[i+8]]
			t2[tcrosswalk]['s1Grade'] = [t2s1[i+9]]
			t2[tcrosswalk]['relationshipID'] = t2s1[i+10]
			t2[tcrosswalk]['s1crosswalked'] = 'true'
		i = i+11

	i = 0;
	for scrosswalk in scrosswalks:
		if 't2URI' in s1[scrosswalk].keys():
			s1[scrosswalk]['t2URI'].append(t2s1[i])
			s1[scrosswalk]['t2Label'].append(t2s1[i+1])
			s1[scrosswalk]['t2Notation'].append(t2s1[i+2])
			s1[scrosswalk]['t2Grade'].append(t2s1[i+3])
		else:
			s1[scrosswalk]['t2URI'] = [t2s1[i]]
			s1[scrosswalk]['t2Label'] = [t2s1[i+1]]
			s1[scrosswalk]['t2Notation'] = [t2s1[i+2]]
			s1[scrosswalk]['t2Grade'] = [t2s1[i+3]]
			s1[scrosswalk]['predicateCURIE'] = t2s1[i+4]
			s1[scrosswalk]['predicateLabel'] = t2s1[i+5]
			s1[scrosswalk]['Label'] = t2s1[i+7]
			s1[scrosswalk]['Notation'] = t2s1[i+8]
			s1[scrosswalk]['Grade'] = t2s1[i+9]
			s1[scrosswalk]['relationshipID'] = t2s1[i+10]
			s1[scrosswalk]['t2crosswalked'] = 'true'
		i = i+11

	j = 0;
	for element in s1:
		s1[element]['id'] = j
		s1[element]['graph'] = 's1'
		j=j+1

	j = 0;
	for element in t1:
		t1[element]['id'] = j
		t1[element]['graph'] = 't1'
		j=j+1

	j = 0;
	for element in t2:
		t2[element]['id'] = j
		t2[element]['graph'] = 't2'
		j=j+1

	t1str = str(t1)
	t1string = ''
	endindex = 0;

	while t1str.find(' u\'') != -1:
		endindex = t1str.find(' u\'')
		t1string = t1string + t1str[:endindex+1]
		t1str = t1str[endindex+2:]

	t1string = t1string + t1str
	t1str = t1string
	t1string = ''
	endindex = 0

	while t1str.find('{u\'') != -1:
		endindex = t1str.find('{u\'')
		t1string = t1string + t1str[:endindex+1]
		t1str = t1str[endindex+2:]

	t1string = t1string + t1str
	t1str = t1string
	t1string = ''
	endindex = 0

	while t1str.find(' u\"') != -1:
		endindex = t1str.find(' u\"')
		t1string = t1string + t1str[:endindex+1]
		t1str = t1str[endindex+2:]

	t1string = t1string + t1str
	t1str = t1string
	t1string = ''
	endindex = 0

	while t1str.find('[u\'') != -1:
		endindex = t1str.find('[u\'')
		t1string = t1string + t1str[:endindex+1]
		t1str = t1str[endindex+2:]

	t1string = t1string + t1str
	t1str = t1string
	t1string = ''
	endindex = 0

	while t1str.find('[u\"') != -1:
		endindex = t1str.find('[u\"')
		t1string = t1string + t1str[:endindex+1]
		t1str = t1str[endindex+2:]

	t1string = t1string + t1str
	t1str = t1string
	t1string = ''
	endindex = 0

	while t1str.find(': {') != -1:
		endindex = t1str.find(': {') + 3
		t1string = t1string + t1str[:endindex] + '\n'
		t1str = t1str[endindex:]

	t1string = t1string + t1str
	t1str = t1string
	t1string = ''
	endindex = 0

	while t1str.find(': [') != -1:
		endindex = t1str.find(': [') + 3
		t1string = t1string + t1str[:endindex] + '\n'
		t1str = t1str[endindex:]

	t1string = t1string + t1str
	t1str = t1string
	t1string = ''
	endindex = 0

	while t1str.find('],') != -1:
		endindex = t1str.find('],') + 2
		t1string = t1string + t1str[:endindex] + '\n'
		t1str = t1str[endindex:]

	t1string = t1string + t1str
	t1str = t1string
	t1string = ''
	endindex = 0

	while t1str.find('},') != -1:
		endindex = t1str.find('},') + 2
		t1string = t1string + t1str[:endindex] + '\n'
		t1str = t1str[endindex:]

	t1string = t1string + t1str
	t1str = t1string
	t1string = ''
	endindex = 0

	while t1str.find(']},') != -1:
		endindex = t1str.find(']},') + 3
		t1string = t1string + t1str[:endindex] + '\n'
		t1str = t1str[endindex:]

	t1string = t1string + t1str
	t1str = t1string
	t1string = ''
	endindex = 0

	while t1str.find('None') != -1:
		endindex = t1str.find('None')
		t1string = t1string + t1str[:endindex] + 'null'
		t1str = t1str[endindex+4:]

	t1string = t1string + t1str



	t2str = str(t2)
	t2string = ''
	endindex = 0;

	while t2str.find(' u\'') != -1:
		endindex = t2str.find(' u\'')
		t2string = t2string + t2str[:endindex+1]
		t2str = t2str[endindex+2:]

	t2string = t2string + t2str
	t2str = t2string
	t2string = ''
	endindex = 0

	while t2str.find('{u\'') != -1:
		endindex = t2str.find('{u\'')
		t2string = t2string + t2str[:endindex+1]
		t2str = t2str[endindex+2:]

	t2string = t2string + t2str
	t2str = t2string
	t2string = ''
	endindex = 0

	while t2str.find(' u\"') != -1:
		endindex = t2str.find(' u\"')
		t2string = t2string + t2str[:endindex+1]
		t2str = t2str[endindex+2:]

	t2string = t2string + t2str
	t2str = t2string
	t2string = ''
	endindex = 0

	while t2str.find('[u\'') != -1:
		endindex = t2str.find('[u\'')
		t2string = t2string + t2str[:endindex+1]
		t2str = t2str[endindex+2:]

	t2string = t2string + t2str
	t2str = t2string
	t2string = ''
	endindex = 0

	while t2str.find('[u\"') != -1:
		endindex = t2str.find('[u\"')
		t2string = t2string + t2str[:endindex+1]
		t2str = t2str[endindex+2:]

	t2string = t2string + t2str
	t2str = t2string
	t2string = ''
	endindex = 0

	while t2str.find(': {') != -1:
		endindex = t2str.find(': {') + 3
		t2string = t2string + t2str[:endindex] + '\n'
		t2str = t2str[endindex:]

	t2string = t2string + t2str
	t2str = t2string
	t2string = ''
	endindex = 0

	while t2str.find(': [') != -1:
		endindex = t2str.find(': [') + 3
		t2string = t2string + t2str[:endindex] + '\n'
		t2str = t2str[endindex:]

	t2string = t2string + t2str
	t2str = t2string
	t2string = ''
	endindex = 0

	while t2str.find('],') != -1:
		endindex = t2str.find('],') + 2
		t2string = t2string + t2str[:endindex] + '\n'
		t2str = t2str[endindex:]

	t2string = t2string + t2str
	t2str = t2string
	t2string = ''
	endindex = 0

	while t2str.find('},') != -1:
		endindex = t2str.find('},') + 2
		t2string = t2string + t2str[:endindex] + '\n'
		t2str = t2str[endindex:]

	t2string = t2string + t2str
	t2str = t2string
	t2string = ''
	endindex = 0

	while t2str.find(']},') != -1:
		endindex = t2str.find(']},') + 3
		t2string = t2string + t2str[:endindex] + '\n'
		t2str = t2str[endindex:]

	t2string = t2string + t2str
	t2str = t2string
	t2string = ''
	endindex = 0

	while t2str.find('None') != -1:
		endindex = t2str.find('None')
		t2string = t2string + t2str[:endindex] + 'null'
		t2str = t2str[endindex+4:]

	t2string = t2string + t2str



	s1str = str(s1)
	s1string = ''
	endindex = 0;

	while s1str.find(' u\'') != -1:
		endindex = s1str.find(' u\'')
		s1string = s1string + s1str[:endindex+1]
		s1str = s1str[endindex+2:]

	s1string = s1string + s1str
	s1str = s1string
	s1string = ''
	endindex = 0

	while s1str.find('{u\'') != -1:
		endindex = s1str.find('{u\'')
		s1string = s1string + s1str[:endindex+1]
		s1str = s1str[endindex+2:]

	s1string = s1string + s1str
	s1str = s1string
	s1string = ''
	endindex = 0

	while s1str.find('[u\'') != -1:
		endindex = s1str.find('[u\'')
		s1string = s1string + s1str[:endindex+1]
		s1str = s1str[endindex+2:]

	s1string = s1string + s1str
	s1str = s1string
	s1string = ''
	endindex = 0

	while s1str.find('[u\"') != -1:
		endindex = s1str.find('[u\"')
		s1string = s1string + s1str[:endindex+1]
		s1str = s1str[endindex+2:]

	s1string = s1string + s1str
	s1str = s1string
	s1string = ''
	endindex = 0

	while s1str.find(' u\"') != -1:
		endindex = s1str.find(' u\"')
		s1string = s1string + s1str[:endindex+1]
		s1str = s1str[endindex+2:]

	s1string = s1string + s1str
	s1str = s1string
	s1string = ''
	endindex = 0

	while s1str.find(': {') != -1:
		endindex = s1str.find(': {') + 3
		s1string = s1string + s1str[:endindex] + '\n'
		s1str = s1str[endindex:]

	s1string = s1string + s1str
	s1str = s1string
	s1string = ''
	endindex = 0

	while s1str.find(': [') != -1:
		endindex = s1str.find(': [') + 3
		s1string = s1string + s1str[:endindex] + '\n'
		s1str = s1str[endindex:]

	s1string = s1string + s1str
	s1str = s1string
	s1string = ''
	endindex = 0

	while s1str.find('],') != -1:
		endindex = s1str.find('],') + 2
		s1string = s1string + s1str[:endindex] + '\n'
		s1str = s1str[endindex:]

	s1string = s1string + s1str
	s1str = s1string
	s1string = ''
	endindex = 0

	while s1str.find('},') != -1:
		endindex = s1str.find('},') + 2
		s1string = s1string + s1str[:endindex] + '\n'
		s1str = s1str[endindex:]

	s1string = s1string + s1str
	s1str = s1string
	s1string = ''
	endindex = 0

	while s1str.find(']},') != -1:
		endindex = s1str.find(']},') + 3
		s1string = s1string + s1str[:endindex] + '\n'
		s1str = s1str[endindex:]

	s1string = s1string + s1str
	s1str = s1string
	s1string = ''
	endindex = 0

	while s1str.find('None') != -1:
		endindex = s1str.find('None')
		s1string = s1string + s1str[:endindex] + 'null'
		s1str = s1str[endindex+4:]

	s1string = s1string + s1str



	output1.write('var s1Data = ' + s1string + ';')
	output1.write('\n\n')
	output1.write('var t1Data = ' + t1string + ';')
	output1.write('\n\n')
	output1.write('var t2Data = ' + t2string + ';')






with open('s2.json') as s2json, open('t3.json') as t3json, open('t4.json') as t4json, open('graph_mindsumo_json_2.js', 'w') as output2:
	s2 = json.load(s2json)
	t3 = json.load(t3json)
	t4 = json.load(t4json)

	sheet3 = openpyxl.load_workbook('t3-s2.xlsx')
	sheet3 = sheet3.get_sheet_by_name('t3-s2')

	sheet4 = openpyxl.load_workbook('t4-s2.xlsx')
	sheet4 = sheet4.get_sheet_by_name('t4-s2')

	t3s2 = []
	tcrosswalks = []
	scrosswalks = []

	for row in range (2, 2740):
		#Find cell names
		su = 'A' + str(row)
		sl = 'B' + str(row)
		sn = 'C' + str(row)
		sg = 'D' + str(row)
		pc = 'E' + str(row)
		pl = 'F' + str(row)
		ou = 'G' + str(row)
		ol = 'H' + str(row)
		on = 'I' + str(row)
		og = 'J' + str(row)
		ri = 'K' + str(row)

		su = sheet3[su].value
		sl = sheet3[sl].value
		sn = sheet3[sn].value
		sg = sheet3[sg].value
		pc = sheet3[pc].value
		pl = sheet3[pl].value
		ou = sheet3[ou].value
		ol = sheet3[ol].value
		on = sheet3[on].value
		og = sheet3[og].value
		ri = sheet3[ri].value

		tcrosswalks.append(su)
		t3s2.append(su)
		t3s2.append(sl)
		t3s2.append(sn)
		t3s2.append(sg)
		t3s2.append(pc)
		t3s2.append(pl)
		scrosswalks.append(ou)
		t3s2.append(ou)
		t3s2.append(ol)
		t3s2.append(on)
		t3s2.append(og)
		t3s2.append(ri)

	i = 0

	for tcrosswalk in tcrosswalks:
		if [t3[tcrosswalk]]:
			if 's2URI' in (t3[tcrosswalk].keys()):
				t3[tcrosswalk]['s2URI'].append(t3s2[i+6])
				t3[tcrosswalk]['s2Label'].append(t3s2[i+7])
				t3[tcrosswalk]['s2Notation'].append(t3s2[i+8])
				t3[tcrosswalk]['s2Grade'].append(t3s2[i+9])
			else:
				t3[tcrosswalk]['Label'] = t3s2[i+1]
				t3[tcrosswalk]['Notation'] = t3s2[i+2]
				t3[tcrosswalk]['Grade'] = t3s2[i+3]
				t3[tcrosswalk]['predicateCURIE'] = t3s2[i+4]
				t3[tcrosswalk]['predicateLabel'] = t3s2[i+5]
				t3[tcrosswalk]['s2URI'] = [t3s2[i+6]]
				t3[tcrosswalk]['s2Label'] = [t3s2[i+7]]
				t3[tcrosswalk]['s2Notation'] = [t3s2[i+8]]
				t3[tcrosswalk]['s2Grade'] = [t3s2[i+9]]
				t3[tcrosswalk]['relationshipID'] = t3s2[i+10]
				t3[tcrosswalk]['s2crosswalked'] = 'true'
		i = i+11

	i = 0;
	for scrosswalk in scrosswalks:
		if 't3URI' in s2[scrosswalk].keys():
			s2[scrosswalk]['t3URI'].append(t3s2[i])
			s2[scrosswalk]['t3Label'].append(t3s2[i+1])
			s2[scrosswalk]['t3Notation'].append(t3s2[i+2])
			s2[scrosswalk]['t3Grade'].append(t3s2[i+3])
		else:
			s2[scrosswalk]['t3URI'] = [t3s2[i]]
			s2[scrosswalk]['t3Label'] = [t3s2[i+1]]
			s2[scrosswalk]['t3Notation'] = [t3s2[i+2]]
			s2[scrosswalk]['t3Grade'] = [t3s2[i+3]]
			s2[scrosswalk]['predicateCURIE'] = t3s2[i+4]
			s2[scrosswalk]['predicateLabel'] = t3s2[i+5]
			s2[scrosswalk]['Label'] = t3s2[i+7]
			s2[scrosswalk]['Notation'] = t3s2[i+8]
			s2[scrosswalk]['Grade'] = t3s2[i+9]
			s2[scrosswalk]['relationshipID'] = t3s2[i+10]
			s2[scrosswalk]['t3crosswalked'] = 'true'
		i = i+11

	t4s2 = []
	tcrosswalks = []
	scrosswalks = []
	for row in range (2, len(sheet4.rows)+1):
		#Find cell names
		su = 'A' + str(row)
		sl = 'B' + str(row)
		sn = 'C' + str(row)
		sg = 'D' + str(row)
		pc = 'E' + str(row)
		pl = 'F' + str(row)
		ou = 'G' + str(row)
		ol = 'H' + str(row)
		on = 'I' + str(row)
		og = 'J' + str(row)
		ri = 'K' + str(row)

		su = sheet4[su].value
		sl = sheet4[sl].value
		sn = sheet4[sn].value
		sg = sheet4[sg].value
		pc = sheet4[pc].value
		pl = sheet4[pl].value
		ou = sheet4[ou].value
		ol = sheet4[ol].value
		on = sheet4[on].value
		og = sheet4[og].value
		ri = sheet4[ri].value

		tcrosswalks.append(su)
		t4s2.append(su)
		t4s2.append(sl)
		t4s2.append(sn)
		t4s2.append(sg)
		t4s2.append(pc)
		t4s2.append(pl)
		scrosswalks.append(ou)
		t4s2.append(ou)
		t4s2.append(ol)
		t4s2.append(on)
		t4s2.append(og)
		t4s2.append(ri)

	i = 0
	for tcrosswalk in tcrosswalks:
		if 's2URI' in t4[tcrosswalk].keys():
			t4[tcrosswalk]['s2URI'].append(t4s2[i+6])
			t4[tcrosswalk]['s2Label'].append(t4s2[i+7])
			t4[tcrosswalk]['s2Notation'].append(t4s2[i+8])
			t4[tcrosswalk]['s2Grade'].append(t4s2[i+9])
		else:
			t4[tcrosswalk]['Label'] = t4s2[i+1]
			t4[tcrosswalk]['Notation'] = t4s2[i+2]
			t4[tcrosswalk]['Grade'] = t4s2[i+3]
			t4[tcrosswalk]['predicateCURIE'] = t4s2[i+4]
			t4[tcrosswalk]['predicateLabel'] = t4s2[i+5]
			t4[tcrosswalk]['s2URI'] = [t4s2[i+6]]
			t4[tcrosswalk]['s2Label'] = [t4s2[i+7]]
			t4[tcrosswalk]['s2Notation'] = [t4s2[i+8]]
			t4[tcrosswalk]['s2Grade'] = [t4s2[i+9]]
			t4[tcrosswalk]['relationshipID'] = t4s2[i+10]
			t4[tcrosswalk]['s2crosswalked'] = 'true'
		i = i+11

	i = 0;
	for scrosswalk in scrosswalks:
		if 't4URI' in s2[scrosswalk].keys():
			s2[scrosswalk]['t4URI'].append(t4s2[i])
			s2[scrosswalk]['t4Label'].append(t4s2[i+1])
			s2[scrosswalk]['t4Notation'].append(t4s2[i+2])
			s2[scrosswalk]['t4Grade'].append(t4s2[i+3])
		else:
			s2[scrosswalk]['t4URI'] = [t4s2[i]]
			s2[scrosswalk]['t4Label'] = [t4s2[i+1]]
			s2[scrosswalk]['t4Notation'] = [t4s2[i+2]]
			s2[scrosswalk]['t4Grade'] = [t4s2[i+3]]
			s2[scrosswalk]['predicateCURIE'] = t4s2[i+4]
			s2[scrosswalk]['predicateLabel'] = t4s2[i+5]
			s2[scrosswalk]['Label'] = t4s2[i+7]
			s2[scrosswalk]['Notation'] = t4s2[i+8]
			s2[scrosswalk]['Grade'] = t4s2[i+9]
			s2[scrosswalk]['relationshipID'] = t4s2[i+10]
			s2[scrosswalk]['t4crosswalked'] = 'true'
		i = i+11

	j = 0;
	for element in s2:
		s2[element]['id'] = j
		s2[element]['graph'] = 's2'
		j=j+1

	j = 0;
	for element in t3:
		t3[element]['id'] = j
		t3[element]['graph'] = 't3'
		j=j+1

	j = 0;
	for element in t4:
		t4[element]['id'] = j
		t4[element]['graph'] = 't4'
		j=j+1

	t3str = str(t3)
	t3string = ''
	endindex = 0;

	while t3str.find(' u\'') != -1:
		endindex = t3str.find(' u\'')
		t3string = t3string + t3str[:endindex+1]
		t3str = t3str[endindex+2:]

	t3string = t3string + t3str
	t3str = t3string
	t3string = ''
	endindex = 0

	while t3str.find('{u\'') != -1:
		endindex = t3str.find('{u\'')
		t3string = t3string + t3str[:endindex+1]
		t3str = t3str[endindex+2:]

	t3string = t3string + t3str
	t3str = t3string
	t3string = ''
	endindex = 0

	while t3str.find(' u\"') != -1:
		endindex = t3str.find(' u\"')
		t3string = t3string + t3str[:endindex+1]
		t3str = t3str[endindex+2:]

	t3string = t3string + t3str
	t3str = t3string
	t3string = ''
	endindex = 0

	while t3str.find('[u\'') != -1:
		endindex = t3str.find('[u\'')
		t3string = t3string + t3str[:endindex+1]
		t3str = t3str[endindex+2:]

	t3string = t3string + t3str
	t3str = t3string
	t3string = ''
	endindex = 0

	while t3str.find('[u\"') != -1:
		endindex = t3str.find('[u\"')
		t3string = t3string + t3str[:endindex+1]
		t3str = t3str[endindex+2:]

	t3string = t3string + t3str
	t3str = t3string
	t3string = ''
	endindex = 0

	while t3str.find(': {') != -1:
		endindex = t3str.find(': {') + 3
		t3string = t3string + t3str[:endindex] + '\n'
		t3str = t3str[endindex:]

	t3string = t3string + t3str
	t3str = t3string
	t3string = ''
	endindex = 0

	while t3str.find(': [') != -1:
		endindex = t3str.find(': [') + 3
		t3string = t3string + t3str[:endindex] + '\n'
		t3str = t3str[endindex:]

	t3string = t3string + t3str
	t3str = t3string
	t3string = ''
	endindex = 0

	while t3str.find('],') != -1:
		endindex = t3str.find('],') + 2
		t3string = t3string + t3str[:endindex] + '\n'
		t3str = t3str[endindex:]

	t3string = t3string + t3str
	t3str = t3string
	t3string = ''
	endindex = 0

	while t3str.find('},') != -1:
		endindex = t3str.find('},') + 2
		t3string = t3string + t3str[:endindex] + '\n'
		t3str = t3str[endindex:]

	t3string = t3string + t3str
	t3str = t3string
	t3string = ''
	endindex = 0

	while t3str.find(']},') != -1:
		endindex = t3str.find(']},') + 3
		t3string = t3string + t3str[:endindex] + '\n'
		t3str = t3str[endindex:]

	t3string = t3string + t3str
	t3str = t3string
	t3string = ''
	endindex = 0

	while t3str.find('None') != -1:
		endindex = t3str.find('None')
		t3string = t3string + t3str[:endindex] + 'null'
		t3str = t3str[endindex+4:]

	t3string = t3string + t3str



	t4str = str(t4)
	t4string = ''
	endindex = 0;

	while t4str.find(' u\'') != -1:
		endindex = t4str.find(' u\'')
		t4string = t4string + t4str[:endindex+1]
		t4str = t4str[endindex+2:]

	t4string = t4string + t4str
	t4str = t4string
	t4string = ''
	endindex = 0

	while t4str.find('{u\'') != -1:
		endindex = t4str.find('{u\'')
		t4string = t4string + t4str[:endindex+1]
		t4str = t4str[endindex+2:]

	t4string = t4string + t4str
	t4str = t4string
	t4string = ''
	endindex = 0

	while t4str.find(' u\"') != -1:
		endindex = t4str.find(' u\"')
		t4string = t4string + t4str[:endindex+1]
		t4str = t4str[endindex+2:]

	t4string = t4string + t4str
	t4str = t4string
	t4string = ''
	endindex = 0

	while t4str.find('[u\'') != -1:
		endindex = t4str.find('[u\'')
		t4string = t4string + t4str[:endindex+1]
		t4str = t4str[endindex+2:]

	t4string = t4string + t4str
	t4str = t4string
	t4string = ''
	endindex = 0

	while t4str.find('[u\"') != -1:
		endindex = t4str.find('[u\"')
		t4string = t4string + t4str[:endindex+1]
		t4str = t4str[endindex+2:]

	t4string = t4string + t4str
	t4str = t4string
	t4string = ''
	endindex = 0

	while t4str.find(': {') != -1:
		endindex = t4str.find(': {') + 3
		t4string = t4string + t4str[:endindex] + '\n'
		t4str = t4str[endindex:]

	t4string = t4string + t4str
	t4str = t4string
	t4string = ''
	endindex = 0

	while t4str.find(': [') != -1:
		endindex = t4str.find(': [') + 3
		t4string = t4string + t4str[:endindex] + '\n'
		t4str = t4str[endindex:]

	t4string = t4string + t4str
	t4str = t4string
	t4string = ''
	endindex = 0

	while t4str.find('],') != -1:
		endindex = t4str.find('],') + 2
		t4string = t4string + t4str[:endindex] + '\n'
		t4str = t4str[endindex:]

	t4string = t4string + t4str
	t4str = t4string
	t4string = ''
	endindex = 0

	while t4str.find('},') != -1:
		endindex = t4str.find('},') + 2
		t4string = t4string + t4str[:endindex] + '\n'
		t4str = t4str[endindex:]

	t4string = t4string + t4str
	t4str = t4string
	t4string = ''
	endindex = 0

	while t4str.find(']},') != -1:
		endindex = t4str.find(']},') + 3
		t4string = t4string + t4str[:endindex] + '\n'
		t4str = t4str[endindex:]

	t4string = t4string + t4str
	t4str = t4string
	t4string = ''
	endindex = 0

	while t4str.find('None') != -1:
		endindex = t4str.find('None')
		t4string = t4string + t4str[:endindex] + 'null'
		t4str = t4str[endindex+4:]

	t4string = t4string + t4str



	s2str = str(s2)
	s2string = ''
	endindex = 0;

	while s2str.find(' u\'') != -1:
		endindex = s2str.find(' u\'')
		s2string = s2string + s2str[:endindex+1]
		s2str = s2str[endindex+2:]

	s2string = s2string + s2str
	s2str = s2string
	s2string = ''
	endindex = 0

	while s2str.find('{u\'') != -1:
		endindex = s2str.find('{u\'')
		s2string = s2string + s2str[:endindex+1]
		s2str = s2str[endindex+2:]

	s2string = s2string + s2str
	s2str = s2string
	s2string = ''
	endindex = 0

	while s2str.find('[u\'') != -1:
		endindex = s2str.find('[u\'')
		s2string = s2string + s2str[:endindex+1]
		s2str = s2str[endindex+2:]

	s2string = s2string + s2str
	s2str = s2string
	s2string = ''
	endindex = 0

	while s2str.find('[u\"') != -1:
		endindex = s2str.find('[u\"')
		s2string = s2string + s2str[:endindex+1]
		s2str = s2str[endindex+2:]

	s2string = s2string + s2str
	s2str = s2string
	s2string = ''
	endindex = 0

	while s2str.find(' u\"') != -1:
		endindex = s2str.find(' u\"')
		s2string = s2string + s2str[:endindex+1]
		s2str = s2str[endindex+2:]

	s2string = s2string + s2str
	s2str = s2string
	s2string = ''
	endindex = 0

	while s2str.find(': {') != -1:
		endindex = s2str.find(': {') + 3
		s2string = s2string + s2str[:endindex] + '\n'
		s2str = s2str[endindex:]

	s2string = s2string + s2str
	s2str = s2string
	s2string = ''
	endindex = 0

	while s2str.find(': [') != -1:
		endindex = s2str.find(': [') + 3
		s2string = s2string + s2str[:endindex] + '\n'
		s2str = s2str[endindex:]

	s2string = s2string + s2str
	s2str = s2string
	s2string = ''
	endindex = 0

	while s2str.find('],') != -1:
		endindex = s2str.find('],') + 2
		s2string = s2string + s2str[:endindex] + '\n'
		s2str = s2str[endindex:]

	s2string = s2string + s2str
	s2str = s2string
	s2string = ''
	endindex = 0

	while s2str.find('},') != -1:
		endindex = s2str.find('},') + 2
		s2string = s2string + s2str[:endindex] + '\n'
		s2str = s2str[endindex:]

	s2string = s2string + s2str
	s2str = s2string
	s2string = ''
	endindex = 0

	while s2str.find(']},') != -1:
		endindex = s2str.find(']},') + 3
		s2string = s2string + s2str[:endindex] + '\n'
		s2str = s2str[endindex:]

	s2string = s2string + s2str
	s2str = s2string
	s2string = ''
	endindex = 0

	while s2str.find('None') != -1:
		endindex = s2str.find('None')
		s2string = s2string + s2str[:endindex] + 'null'
		s2str = s2str[endindex+4:]

	s2string = s2string + s2str

	output2.write('var s2Data = ' + s2string + ';')
	output2.write('\n\n')
	output2.write('var t3Data = ' + t3string + ';')
	output2.write('\n\n')
	output2.write('var t4Data = ' + t4string + ';')
